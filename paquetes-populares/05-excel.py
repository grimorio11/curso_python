import openpyxl 

# Cargar el libro de trabajo
wb = openpyxl.load_workbook('Planilla.xlsx')

print(wb.sheetnames)  # Imprimir los nombres de las hojas
print(wb["Hoja1"])  # Imprimir la hoja 1
print(wb["Hoja1"].title)  # Imprimir el título de la hoja 1

hoja = wb.active  # Obtener la hoja activa

wb.create_sheet("curso")  # Crear una nueva hoja llamada "curso"
print(wb.sheetnames)

hoja1 = wb["Hoja1"]  # Obtener la hoja "Hoja1"
print(hoja1)
print(
    hoja1.max_row,
    hoja1.max_column,
)

# Imprimir el valor de la celda A1
celda = hoja1["A1"]
celda.value = "Hola Mundo"  # Cambiar el valor de la celda A1
print(celda)  # Imprimir la celda A1
print(celda.value)  # Imprimir el valor de la celda A1

celda2 = hoja1.cell(row=2, column=1)  # Obtener la celda B2
print(celda2.value)  # Imprimir la celda B2
print(
    celda2.value,
    celda2.row,
    celda2.column,
    celda2.coordinate,
)

for fila in range(1, hoja1.max_row + 1):
    for columna in range(1, hoja1.max_column + 1):
        celda = hoja1.cell(row=fila, column=columna)
        print(fila, columna, celda.value)

columna = hoja1["A"]  # Obtener la columna A
fila = hoja1["1"]  # Obtener la fila 1
print(columna)  # Imprimir la columna A
print(fila)  # Imprimir la fila 1

hoja1.append(["Hola", "27", "hola@chanchito.com"])  # Agregar una fila al final de la hoja
print(hoja1.rows)  # Imprimir todas las filas de la hoja

hoja1.delete_rows(2, 1)  # Eliminar la fila 2
hoja1.delete_cols(2, 1)  # Eliminar la columna 2

wb.save("Planilla2.xlsx")  # Guardar los cambios en el libro de trabajo
#wb.close()  # Cerrar el libro de trabajo

 